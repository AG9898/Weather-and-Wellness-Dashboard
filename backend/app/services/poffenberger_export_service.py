from __future__ import annotations

import io
import json
from collections.abc import Mapping
from datetime import date, datetime, timedelta, timezone
from decimal import Decimal
from typing import Any

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.participants import Participant
from app.models.poffenberger import PoffenbergerRun
from app.models.sessions import Session as SessionModel


_DATA_COLUMNS = [
    "participant_number",
    "age_band",
    "gender",
    "handedness",
    "trial_time",
    "lh_lvf_accuracy",
    "lh_lvf_mean_rt_ms",
    "lh_rvf_accuracy",
    "lh_rvf_mean_rt_ms",
    "rh_lvf_accuracy",
    "rh_lvf_mean_rt_ms",
    "rh_rvf_accuracy",
    "rh_rvf_mean_rt_ms",
    "mean_rt_crossed_ms",
    "mean_rt_uncrossed_ms",
    "ihtt_difference_ms",
    "accuracy_crossed",
    "accuracy_uncrossed",
]

_COLUMN_LABELS = {
    "participant_number": "Participant number",
    "age_band": "Age group",
    "gender": "Gender",
    "handedness": "Handedness",
    "trial_time": "Trial Time",
    "lh_lvf_accuracy": "Left hand + left-side stimulus accuracy",
    "lh_lvf_mean_rt_ms": "Left hand + left-side stimulus mean RT (ms)",
    "lh_rvf_accuracy": "Left hand + right-side stimulus accuracy",
    "lh_rvf_mean_rt_ms": "Left hand + right-side stimulus mean RT (ms)",
    "rh_lvf_accuracy": "Right hand + left-side stimulus accuracy",
    "rh_lvf_mean_rt_ms": "Right hand + left-side stimulus mean RT (ms)",
    "rh_rvf_accuracy": "Right hand + right-side stimulus accuracy",
    "rh_rvf_mean_rt_ms": "Right hand + right-side stimulus mean RT (ms)",
    "mean_rt_crossed_ms": "Crossed mean RT (ms)",
    "mean_rt_uncrossed_ms": "Uncrossed mean RT (ms)",
    "ihtt_difference_ms": "IHTT difference (ms)",
    "accuracy_crossed": "Crossed accuracy",
    "accuracy_uncrossed": "Uncrossed accuracy",
}

_COLUMN_WIDTHS = {
    "participant_number": 20,
    "age_band": 16,
    "gender": 18,
    "handedness": 20,
    "trial_time": 36,
    "mean_rt_crossed_ms": 22,
    "mean_rt_uncrossed_ms": 24,
    "ihtt_difference_ms": 20,
    "accuracy_crossed": 18,
    "accuracy_uncrossed": 20,
}

_README_ROWS = [
    ("IHTT Poffenberger - Data Export", "title"),
    ("Exported", "field"),
    ("Workbook type", "field"),
    (None, "blank"),
    ("How to use this workbook", "heading"),
    ("Open the Poffenberger Data sheet. Each row is one recorded participant session.", "body"),
    (
        "The columns combine start-of-session demographics with the task results that summarize the participant's run.",
        "body",
    ),
    (
        "Blank result cells usually mean the session was not submitted or did not have enough valid responses for that score.",
        "body",
    ),
    (None, "blank"),
    ("Score notes", "heading"),
    ("Reaction-time columns are in milliseconds.", "body"),
    ("Accuracy columns are formatted as percentages.", "body"),
    (
        "IHTT difference is crossed mean reaction time minus uncrossed mean reaction time.",
        "body",
    ),
    (
        "The four hand/stimulus columns show the main condition summaries used to review the run.",
        "body",
    ),
]

_PAPER_FILL = PatternFill("solid", fgColor="FBFAF6")
_CARD_FILL = PatternFill("solid", fgColor="FFFFFF")
_HEADER_FILL = PatternFill("solid", fgColor="001328")
_SECTION_FILL = PatternFill("solid", fgColor="CFD7DE")
_ALT_ROW_FILL = PatternFill("solid", fgColor="F4F1EA")
_HAIRLINE = Side(style="thin", color="D9D5CC")
_BORDER = Border(left=_HAIRLINE, right=_HAIRLINE, top=_HAIRLINE, bottom=_HAIRLINE)
_FONT_NAME = "JetBrains Mono"
_PERCENT_COLUMNS = {
    "lh_lvf_accuracy",
    "lh_rvf_accuracy",
    "rh_lvf_accuracy",
    "rh_rvf_accuracy",
    "accuracy_crossed",
    "accuracy_uncrossed",
}
_MILLISECONDS_COLUMNS = {
    "lh_lvf_mean_rt_ms",
    "lh_rvf_mean_rt_ms",
    "rh_lvf_mean_rt_ms",
    "rh_rvf_mean_rt_ms",
    "mean_rt_crossed_ms",
    "mean_rt_uncrossed_ms",
    "ihtt_difference_ms",
}

_RUN_SUMMARY_COLUMNS = [
    "lh_lvf_accuracy",
    "lh_lvf_mean_rt_ms",
    "lh_rvf_accuracy",
    "lh_rvf_mean_rt_ms",
    "rh_lvf_accuracy",
    "rh_lvf_mean_rt_ms",
    "rh_rvf_accuracy",
    "rh_rvf_mean_rt_ms",
    "mean_rt_crossed_ms",
    "mean_rt_uncrossed_ms",
    "ihtt_difference_ms",
    "accuracy_crossed",
    "accuracy_uncrossed",
]

_SAMPLE_EXPORTED_AT = datetime(2026, 6, 24, 17, 30, tzinfo=timezone.utc)


def _to_xlsx(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, Decimal):
        return float(value)
    if isinstance(value, (dict, list)):
        return json.dumps(value)
    return value


def _format_trial_time(started_at: Any, completed_at: Any) -> str | None:
    started = _to_xlsx(started_at)
    completed = _to_xlsx(completed_at)
    if started is None and completed is None:
        return None
    return f"Started: {started or ''}\nCompleted: {completed or ''}"


def _data_cell_value(row: Mapping[str, Any], column: str) -> Any:
    if column == "trial_time":
        return _format_trial_time(row.get("started_at"), row.get("completed_at"))
    return _to_xlsx(row[column])


async def _fetch_run_rows(db: AsyncSession) -> list[dict[str, Any]]:
    stmt = (
        select(
            Participant.participant_number.label("participant_number"),
            Participant.age_band.label("age_band"),
            Participant.gender.label("gender"),
            Participant.handedness.label("handedness"),
            PoffenbergerRun.started_at.label("started_at"),
            PoffenbergerRun.completed_at.label("completed_at"),
            *[
                getattr(PoffenbergerRun, column).label(column)
                for column in _RUN_SUMMARY_COLUMNS
            ],
        )
        .join(
            Participant,
            Participant.participant_uuid == PoffenbergerRun.participant_uuid,
        )
        .join(SessionModel, SessionModel.session_id == PoffenbergerRun.session_id)
        .order_by(PoffenbergerRun.started_at, PoffenbergerRun.run_id)
    )
    return list((await db.execute(stmt)).mappings().all())


def _sample_run_rows() -> list[dict[str, Any]]:
    samples = [
        (
            9001,
            "18-24",
            "Woman",
            "Right-handed",
            "0.9467",
            "318.40",
            "0.9600",
            "325.80",
            "0.9533",
            "329.20",
            "0.9733",
            "316.60",
            "327.50",
            "317.50",
            "10.00",
            "0.9567",
            "0.9600",
        ),
        (
            9002,
            "25-31",
            "Man",
            "Left-handed",
            "0.9733",
            "302.15",
            "0.9533",
            "309.90",
            "0.9600",
            "314.45",
            "0.9800",
            "300.85",
            "312.18",
            "301.50",
            "10.68",
            "0.9567",
            "0.9767",
        ),
        (
            9003,
            "32-38",
            "Non-binary",
            "Ambidextrous",
            "0.9200",
            "342.30",
            "0.9333",
            "351.10",
            "0.9267",
            "347.75",
            "0.9467",
            "337.90",
            "349.42",
            "340.10",
            "9.32",
            "0.9300",
            "0.9333",
        ),
        (
            9004,
            ">38",
            "Prefer not to say",
            "Prefer not to say",
            "0.9867",
            "289.65",
            "0.9733",
            "298.20",
            "0.9800",
            "296.40",
            "0.9933",
            "287.35",
            "297.30",
            "288.50",
            "8.80",
            "0.9767",
            "0.9900",
        ),
    ]
    rows: list[dict[str, Any]] = []
    for index, sample in enumerate(samples):
        (
            participant_number,
            age_band,
            gender,
            handedness,
            lh_lvf_accuracy,
            lh_lvf_mean_rt_ms,
            lh_rvf_accuracy,
            lh_rvf_mean_rt_ms,
            rh_lvf_accuracy,
            rh_lvf_mean_rt_ms,
            rh_rvf_accuracy,
            rh_rvf_mean_rt_ms,
            mean_rt_crossed_ms,
            mean_rt_uncrossed_ms,
            ihtt_difference_ms,
            accuracy_crossed,
            accuracy_uncrossed,
        ) = sample
        started_at = _SAMPLE_EXPORTED_AT + timedelta(minutes=index * 22)
        rows.append(
            {
                "participant_number": participant_number,
                "age_band": age_band,
                "gender": gender,
                "handedness": handedness,
                "started_at": started_at,
                "completed_at": started_at + timedelta(minutes=18),
                "lh_lvf_accuracy": Decimal(lh_lvf_accuracy),
                "lh_lvf_mean_rt_ms": Decimal(lh_lvf_mean_rt_ms),
                "lh_rvf_accuracy": Decimal(lh_rvf_accuracy),
                "lh_rvf_mean_rt_ms": Decimal(lh_rvf_mean_rt_ms),
                "rh_lvf_accuracy": Decimal(rh_lvf_accuracy),
                "rh_lvf_mean_rt_ms": Decimal(rh_lvf_mean_rt_ms),
                "rh_rvf_accuracy": Decimal(rh_rvf_accuracy),
                "rh_rvf_mean_rt_ms": Decimal(rh_rvf_mean_rt_ms),
                "mean_rt_crossed_ms": Decimal(mean_rt_crossed_ms),
                "mean_rt_uncrossed_ms": Decimal(mean_rt_uncrossed_ms),
                "ihtt_difference_ms": Decimal(ihtt_difference_ms),
                "accuracy_crossed": Decimal(accuracy_crossed),
                "accuracy_uncrossed": Decimal(accuracy_uncrossed),
            }
        )
    return rows


def _style_sheet_background(ws: Any, rows: int, columns: int) -> None:
    for row in ws.iter_rows(min_row=1, max_row=rows, min_col=1, max_col=columns):
        for cell in row:
            cell.fill = _PAPER_FILL
            cell.font = Font(name=_FONT_NAME, size=10, color="1F2933")


def _write_readme(ws: Any, *, export_date: str, export_mode: str) -> None:
    ws.title = "README"
    _style_sheet_background(ws, rows=24, columns=3)
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 96

    row_idx = 1
    for text, kind in _README_ROWS:
        if kind == "blank":
            row_idx += 1
            continue
        if kind == "field":
            value = export_date if text == "Exported" else export_mode
            label_cell = ws.cell(row=row_idx, column=1, value=text)
            value_cell = ws.cell(row=row_idx, column=2, value=value)
            for cell in (label_cell, value_cell):
                cell.fill = _CARD_FILL
                cell.border = _BORDER
                cell.alignment = Alignment(vertical="top", wrap_text=True)
            label_cell.font = Font(name=_FONT_NAME, size=10, bold=True, color="001328")
            value_cell.font = Font(name=_FONT_NAME, size=10, color="1F2933")
        else:
            cell = ws.cell(row=row_idx, column=1, value=text)
            ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=2)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if kind == "title":
                cell.fill = _HEADER_FILL
                cell.font = Font(name=_FONT_NAME, size=14, bold=True, color="FFFFFF")
            elif kind == "heading":
                cell.fill = _SECTION_FILL
                cell.font = Font(name=_FONT_NAME, size=11, bold=True, color="001328")
            else:
                cell.fill = _CARD_FILL
                cell.font = Font(name=_FONT_NAME, size=10, color="1F2933")
            cell.border = _BORDER
        row_idx += 1


def _write_data_sheet(ws: Any, rows: list[Mapping[str, Any]]) -> None:
    ws.title = "Poffenberger Data"
    _style_sheet_background(ws, rows=max(8, len(rows) + 6), columns=len(_DATA_COLUMNS))

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(_DATA_COLUMNS))
    title = ws.cell(row=1, column=1, value="Poffenberger participant and trial summary")
    title.fill = _HEADER_FILL
    title.font = Font(name=_FONT_NAME, size=13, bold=True, color="FFFFFF")
    title.alignment = Alignment(vertical="center")
    title.border = _BORDER

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(_DATA_COLUMNS))
    note = ws.cell(
        row=2,
        column=1,
        value=(
            "One row per recorded participant session. Demographics and the main "
            "reaction-time/accuracy summaries are kept together for review."
        ),
    )
    note.fill = _CARD_FILL
    note.font = Font(name=_FONT_NAME, size=10, color="1F2933")
    note.alignment = Alignment(vertical="center", wrap_text=True)
    note.border = _BORDER

    header_row = 4
    for col_idx, column in enumerate(_DATA_COLUMNS, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=_COLUMN_LABELS[column])
        cell.fill = _HEADER_FILL
        cell.font = Font(name=_FONT_NAME, size=10, bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _BORDER
        letter = get_column_letter(col_idx)
        ws.column_dimensions[letter].width = _COLUMN_WIDTHS.get(column, 32)

    for row_idx, row in enumerate(rows, start=header_row + 1):
        row_fill = _CARD_FILL if row_idx % 2 else _ALT_ROW_FILL
        for col_idx, column in enumerate(_DATA_COLUMNS, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=_data_cell_value(row, column))
            cell.fill = row_fill
            cell.border = _BORDER
            cell.font = Font(name=_FONT_NAME, size=10, color="1F2933")
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if column in _PERCENT_COLUMNS:
                cell.number_format = "0.0%"
            elif column in _MILLISECONDS_COLUMNS:
                cell.number_format = "0.00"
            elif column == "trial_time":
                cell.number_format = "@"

    ws.freeze_panes = "A5"
    ws.auto_filter.ref = (
        f"A{header_row}:{get_column_letter(len(_DATA_COLUMNS))}{max(header_row + 1, header_row + len(rows))}"
    )
    ws.sheet_view.showGridLines = False


def _build_workbook(
    *,
    export_date: str,
    export_mode: str,
    run_rows: list[Mapping[str, Any]],
) -> bytes:
    wb = openpyxl.Workbook()

    readme_ws = wb.active
    _write_readme(readme_ws, export_date=export_date, export_mode=export_mode)

    data_ws = wb.create_sheet(title="Poffenberger Data")
    _write_data_sheet(data_ws, run_rows)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


async def build_poffenberger_xlsx(db: AsyncSession, export_date: str) -> bytes:
    run_rows = await _fetch_run_rows(db)
    return _build_workbook(
        export_date=export_date,
        export_mode="Recorded study data export",
        run_rows=run_rows,
    )


def build_sample_poffenberger_xlsx(export_date: str) -> bytes:
    return _build_workbook(
        export_date=export_date,
        export_mode="Sample workbook with fictional rows",
        run_rows=_sample_run_rows(),
    )
