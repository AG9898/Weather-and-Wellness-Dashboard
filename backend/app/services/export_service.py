"""Export service: generate XLSX workbook and zipped CSV bundle from all DB tables.

Tables exported (in order):
  participants, sessions, survey_uls8, survey_cesd10, survey_gad7,
  survey_cogfunc8a, digitspan_runs, digitspan_trials, study_days,
  weather_ingest_runs, weather_daily, imported_session_measures

All exports are schema-faithful and include join keys (participant_uuid,
session_id, study_day_id, etc.) needed to link tables for analysis.

Value conventions:
- UUIDs → string
- datetime → ISO-8601 string (UTC)
- date → ISO-8601 string
- JSONB (dict/list) → JSON string
- bool → Python bool (XLSX) / "true"/"false" (CSV)
- None → empty cell (XLSX) / empty string (CSV)
"""
from __future__ import annotations

import csv
import io
import json
import uuid
import zipfile
from dataclasses import dataclass
from datetime import date, datetime
from typing import Any, Type

import openpyxl
from openpyxl.styles import Font
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.digitspan import DigitSpanRun, DigitSpanTrial
from app.models.imported_session_measures import ImportedSessionMeasures
from app.models.participants import Participant
from app.models.sessions import Session as SessionModel
from app.models.surveys import SurveyCESD10, SurveyCogFunc8a, SurveyGAD7, SurveyULS8
from app.models.weather import StudyDay, WeatherDaily, WeatherIngestRun


# ── Table specifications ────────────────────────────────────────────────────────

@dataclass
class _TableSpec:
    name: str           # Table / sheet / CSV filename stem
    description: str    # One-line description for the README sheet
    model: Any          # SQLAlchemy model class
    columns: list[str]  # Column names in export order
    order_by: list[str] # Column name(s) for ORDER BY


_TABLE_SPECS: list[_TableSpec] = [
    _TableSpec(
        name="participants",
        description="One row per participant (anonymous). Join key: participant_uuid.",
        model=Participant,
        columns=[
            "participant_uuid", "participant_number", "created_at",
            "age_band", "gender", "origin", "origin_other_text",
            "commute_method", "commute_method_other_text", "time_outside",
            "daylight_exposure_minutes",
        ],
        order_by=["participant_number"],
    ),
    _TableSpec(
        name="sessions",
        description="One row per session. Join keys: session_id, participant_uuid, study_day_id.",
        model=SessionModel,
        columns=[
            "session_id", "participant_uuid", "status",
            "created_at", "completed_at", "study_day_id",
        ],
        order_by=["created_at"],
    ),
    _TableSpec(
        name="survey_uls8",
        description="ULS-8 loneliness survey responses. Join keys: session_id, participant_uuid.",
        model=SurveyULS8,
        columns=[
            "response_id", "session_id", "participant_uuid",
            "r1", "r2", "r3", "r4", "r5", "r6", "r7", "r8",
            "computed_mean", "score_0_100", "created_at",
        ],
        order_by=["created_at"],
    ),
    _TableSpec(
        name="survey_cesd10",
        description="CES-D 10 depression survey responses. Join keys: session_id, participant_uuid.",
        model=SurveyCESD10,
        columns=[
            "response_id", "session_id", "participant_uuid",
            "r1", "r2", "r3", "r4", "r5", "r6", "r7", "r8", "r9", "r10",
            "total_score", "created_at",
        ],
        order_by=["created_at"],
    ),
    _TableSpec(
        name="survey_gad7",
        description="GAD-7 anxiety survey responses. Join keys: session_id, participant_uuid.",
        model=SurveyGAD7,
        columns=[
            "response_id", "session_id", "participant_uuid",
            "r1", "r2", "r3", "r4", "r5", "r6", "r7",
            "total_score", "severity_band", "created_at",
        ],
        order_by=["created_at"],
    ),
    _TableSpec(
        name="survey_cogfunc8a",
        description=(
            "CogFunc 8a cognitive function survey rows "
            "(native responses or imported legacy aggregates). "
            "Join keys: session_id, participant_uuid."
        ),
        model=SurveyCogFunc8a,
        columns=[
            "response_id", "session_id", "participant_uuid",
            "r1", "r2", "r3", "r4", "r5", "r6", "r7", "r8",
            "total_sum", "mean_score", "legacy_mean_1_5",
            "data_source", "created_at",
        ],
        order_by=["created_at"],
    ),
    _TableSpec(
        name="digitspan_runs",
        description="Backwards Digit Span run summaries. Join keys: run_id, session_id, participant_uuid.",
        model=DigitSpanRun,
        columns=[
            "run_id", "session_id", "participant_uuid",
            "total_correct", "max_span", "created_at",
        ],
        order_by=["created_at"],
    ),
    _TableSpec(
        name="digitspan_trials",
        description="Individual Digit Span trials (14 per run). Join key: run_id.",
        model=DigitSpanTrial,
        columns=[
            "trial_id", "run_id",
            "trial_number", "span_length",
            "sequence_shown", "sequence_entered", "correct",
        ],
        order_by=["run_id", "trial_number"],
    ),
    _TableSpec(
        name="study_days",
        description="Day dimension table (one row per local calendar day). Join key: study_day_id.",
        model=StudyDay,
        columns=["study_day_id", "date_local", "tz_name", "created_at"],
        order_by=["date_local"],
    ),
    _TableSpec(
        name="weather_ingest_runs",
        description="Audit record for each weather ingestion attempt. Join key: run_id.",
        model=WeatherIngestRun,
        columns=[
            "run_id", "station_id", "date_local", "ingested_at",
            "requested_via", "requested_by_lab_member_id",
            "source_primary_url", "source_secondary_url",
            "http_primary_status", "http_secondary_status",
            "raw_html_primary", "raw_html_secondary",
            "raw_html_primary_sha256", "raw_html_secondary_sha256",
            "parsed_json", "parse_status", "parse_errors",
            "parser_version", "created_at",
        ],
        order_by=["ingested_at"],
    ),
    _TableSpec(
        name="weather_daily",
        description="Day-level weather summary per station. Join keys: study_day_id, source_run_id.",
        model=WeatherDaily,
        columns=[
            "daily_id", "station_id", "study_day_id", "date_local",
            "source_run_id", "updated_at", "current_observed_at",
            "current_temp_c", "current_relative_humidity_pct",
            "current_wind_speed_kmh", "current_wind_gust_kmh",
            "current_wind_dir_deg", "current_pressure_kpa",
            "current_precip_today_mm",
            "forecast_high_c", "forecast_low_c",
            "forecast_precip_prob_pct", "forecast_precip_mm",
            "forecast_condition_text", "forecast_periods",
            "structured_json", "created_at",
        ],
        order_by=["date_local", "station_id"],
    ),
    _TableSpec(
        name="imported_session_measures",
        description="Legacy aggregate outcomes from imported data. Join keys: session_id, participant_uuid.",
        model=ImportedSessionMeasures,
        columns=[
            "session_id", "participant_uuid",
            "precipitation_mm", "temperature_c",
            "anxiety_mean", "loneliness_mean", "depression_mean",
            "digit_span_max_span", "self_report",
            "source_row_json", "created_at",
        ],
        order_by=["created_at"],
    ),
]


# ── Value converters ───────────────────────────────────────────────────────────

def _to_xlsx(value: Any) -> Any:
    """Convert a DB value to an XLSX-compatible Python value.

    UUIDs, dates, and datetimes become ISO strings to avoid timezone and
    format complications.  JSONB (dict/list) become JSON strings.
    Numbers and booleans are kept as-is for Excel analysis.
    """
    if value is None:
        return None
    if isinstance(value, uuid.UUID):
        return str(value)
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, (dict, list)):
        return json.dumps(value)
    return value


def _to_csv(value: Any) -> str:
    """Convert a DB value to a CSV-safe string."""
    if value is None:
        return ""
    if isinstance(value, uuid.UUID):
        return str(value)
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, (dict, list)):
        return json.dumps(value)
    if isinstance(value, bool):
        return "true" if value else "false"
    return str(value)


# ── DB fetch ───────────────────────────────────────────────────────────────────

async def _fetch_rows(
    db: AsyncSession,
    spec: _TableSpec,
) -> list[list[Any]]:
    """Fetch all rows for a table spec, ordered as configured.

    Returns a list of rows; each row is a list of values in column order.
    """
    stmt = select(spec.model)
    for col_name in spec.order_by:
        col = getattr(spec.model, col_name)
        stmt = stmt.order_by(col)

    result = await db.execute(stmt)
    orm_rows = result.scalars().all()

    return [
        [getattr(row, col) for col in spec.columns]
        for row in orm_rows
    ]


# ── XLSX builder ───────────────────────────────────────────────────────────────

_README_TEXT = """\
Weather & Wellness + Misokinesia Research — Data Export

This workbook contains all data from the lab web application database.
One sheet per table; each sheet uses the canonical column order from the schema.

Tables included:
{table_list}

Join key conventions:
  - participant_uuid : links participants → sessions → survey_* → digitspan_runs → imported_session_measures
  - session_id       : links sessions → survey_* → digitspan_runs → digitspan_trials → imported_session_measures
  - run_id           : links digitspan_runs → digitspan_trials
  - study_day_id     : links study_days → sessions → weather_daily
  - source_run_id    : links weather_ingest_runs → weather_daily

Value notes:
  - All UUIDs are strings.
  - Timestamps are ISO-8601 UTC strings.
  - JSONB columns (parsed_json, forecast_periods, source_row_json, etc.) contain JSON strings.
  - raw_html_primary / raw_html_secondary may contain large HTML payloads (debug data).
"""


async def build_xlsx(db: AsyncSession, export_date: str) -> bytes:
    """Fetch all DB tables and return a schema-faithful XLSX workbook as bytes.

    Args:
        db: Async DB session.
        export_date: Today's local date string (YYYY-MM-DD) used in README.

    Returns:
        Raw XLSX bytes suitable for streaming.
    """
    wb = openpyxl.Workbook()

    # ── README sheet ─────────────────────────────────────────────────────────
    readme_ws = wb.active
    readme_ws.title = "README"

    table_list_lines = "\n".join(
        f"  {spec.name:<30} {spec.description}"
        for spec in _TABLE_SPECS
    )
    readme_text = _README_TEXT.format(table_list=table_list_lines)

    for i, line in enumerate(readme_text.splitlines(), start=1):
        cell = readme_ws.cell(row=i, column=1, value=line)
        if i == 1:
            cell.font = Font(bold=True, size=13)

    readme_ws.cell(row=2, column=1, value=f"Exported: {export_date}")
    readme_ws.column_dimensions["A"].width = 100

    # ── One sheet per table ───────────────────────────────────────────────────
    header_font = Font(bold=True)

    for spec in _TABLE_SPECS:
        ws = wb.create_sheet(title=spec.name[:31])  # Excel limit: 31 chars

        # Header row
        for col_idx, col_name in enumerate(spec.columns, start=1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = header_font

        ws.freeze_panes = "A2"

        # Data rows
        data_rows = await _fetch_rows(db, spec)
        for row_idx, row_values in enumerate(data_rows, start=2):
            for col_idx, value in enumerate(row_values, start=1):
                ws.cell(row=row_idx, column=col_idx, value=_to_xlsx(value))

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


# ── ZIP CSV builder ────────────────────────────────────────────────────────────

async def build_zip_csv(db: AsyncSession) -> bytes:
    """Fetch all DB tables and return a ZIP of CSVs as bytes.

    The ZIP contains one CSV per table named '<table_name>.csv'.

    Returns:
        Raw ZIP bytes suitable for streaming.
    """
    buffer = io.BytesIO()

    with zipfile.ZipFile(buffer, mode="w", compression=zipfile.ZIP_DEFLATED) as zf:
        for spec in _TABLE_SPECS:
            csv_buffer = io.StringIO()
            writer = csv.writer(csv_buffer, lineterminator="\n")

            # Header
            writer.writerow(spec.columns)

            # Data
            data_rows = await _fetch_rows(db, spec)
            for row_values in data_rows:
                writer.writerow([_to_csv(v) for v in row_values])

            zf.writestr(f"{spec.name}.csv", csv_buffer.getvalue())

    buffer.seek(0)
    return buffer.getvalue()
