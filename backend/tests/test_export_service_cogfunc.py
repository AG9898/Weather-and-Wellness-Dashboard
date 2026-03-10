"""Regression tests for CogFunc export parity after imported-row support."""

from __future__ import annotations

import csv
import io
import uuid
import zipfile
from datetime import datetime, timezone
from types import SimpleNamespace
from unittest import IsolatedAsyncioTestCase

import openpyxl

from app.services import export_service


class _FakeResult:
    def __init__(self, rows: list[object] | None = None) -> None:
        self._rows = rows or []

    def scalars(self) -> "_FakeResult":
        return self

    def all(self) -> list[object]:
        return list(self._rows)


class _FakeAsyncSession:
    def __init__(self, execute_results: list[_FakeResult]) -> None:
        self.execute_results = list(execute_results)

    async def execute(self, stmt: object) -> _FakeResult:
        if not self.execute_results:
            raise AssertionError("Unexpected execute() call with no remaining fake results.")
        return self.execute_results.pop(0)


def _session_for_cogfunc_row(cogfunc_row: object) -> _FakeAsyncSession:
    results: list[_FakeResult] = []
    for spec in export_service._TABLE_SPECS:
        rows = [cogfunc_row] if spec.name == "survey_cogfunc8a" else []
        results.append(_FakeResult(rows))
    return _FakeAsyncSession(results)


class ExportServiceCogFuncTests(IsolatedAsyncioTestCase):
    async def test_build_xlsx_includes_imported_cogfunc_columns_and_values(self) -> None:
        created_at = datetime(2026, 3, 10, 18, 45, tzinfo=timezone.utc)
        row = SimpleNamespace(
            response_id=uuid.uuid4(),
            session_id=uuid.uuid4(),
            participant_uuid=uuid.uuid4(),
            r1=None,
            r2=None,
            r3=None,
            r4=None,
            r5=None,
            r6=None,
            r7=None,
            r8=None,
            total_sum=None,
            mean_score=None,
            legacy_mean_1_5=4.25,
            data_source="imported",
            created_at=created_at,
        )

        workbook_bytes = await export_service.build_xlsx(
            _session_for_cogfunc_row(row),
            export_date="2026-03-10",
        )

        workbook = openpyxl.load_workbook(io.BytesIO(workbook_bytes))
        sheet = workbook["survey_cogfunc8a"]
        headers = [cell.value for cell in sheet[1]]

        self.assertEqual(
            headers,
            [
                "response_id",
                "session_id",
                "participant_uuid",
                "r1",
                "r2",
                "r3",
                "r4",
                "r5",
                "r6",
                "r7",
                "r8",
                "total_sum",
                "mean_score",
                "legacy_mean_1_5",
                "data_source",
                "created_at",
            ],
        )

        values = {
            headers[index]: sheet.cell(row=2, column=index + 1).value
            for index in range(len(headers))
        }
        self.assertEqual(values["legacy_mean_1_5"], 4.25)
        self.assertEqual(values["data_source"], "imported")
        self.assertEqual(values["created_at"], created_at.isoformat())
        self.assertIsNone(values["mean_score"])

        readme_values = [
            workbook["README"].cell(row=row_num, column=1).value
            for row_num in range(1, 30)
        ]
        self.assertTrue(
            any(
                value
                and "survey_cogfunc8a" in str(value)
                and "imported legacy aggregates" in str(value)
                for value in readme_values
            )
        )

    async def test_build_zip_csv_includes_imported_cogfunc_columns_and_values(self) -> None:
        created_at = datetime(2026, 3, 10, 18, 45, tzinfo=timezone.utc)
        row = SimpleNamespace(
            response_id=uuid.uuid4(),
            session_id=uuid.uuid4(),
            participant_uuid=uuid.uuid4(),
            r1=None,
            r2=None,
            r3=None,
            r4=None,
            r5=None,
            r6=None,
            r7=None,
            r8=None,
            total_sum=None,
            mean_score=None,
            legacy_mean_1_5=3.5,
            data_source="imported",
            created_at=created_at,
        )

        archive_bytes = await export_service.build_zip_csv(_session_for_cogfunc_row(row))

        with zipfile.ZipFile(io.BytesIO(archive_bytes)) as archive:
            with archive.open("survey_cogfunc8a.csv") as csv_file:
                decoded = io.TextIOWrapper(csv_file, encoding="utf-8")
                rows = list(csv.reader(decoded))

        self.assertEqual(
            rows[0],
            [
                "response_id",
                "session_id",
                "participant_uuid",
                "r1",
                "r2",
                "r3",
                "r4",
                "r5",
                "r6",
                "r7",
                "r8",
                "total_sum",
                "mean_score",
                "legacy_mean_1_5",
                "data_source",
                "created_at",
            ],
        )

        row_values = dict(zip(rows[0], rows[1], strict=True))
        self.assertEqual(row_values["legacy_mean_1_5"], "3.5")
        self.assertEqual(row_values["data_source"], "imported")
        self.assertEqual(row_values["created_at"], created_at.isoformat())
        self.assertEqual(row_values["mean_score"], "")
