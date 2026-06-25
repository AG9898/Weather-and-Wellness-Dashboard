from __future__ import annotations

import io
import uuid
from datetime import datetime, timezone
from decimal import Decimal
from types import SimpleNamespace
from unittest import IsolatedAsyncioTestCase

import openpyxl

from app.services.poffenberger_export_service import (
    build_poffenberger_xlsx,
    build_sample_poffenberger_xlsx,
)


class _MappingRows:
    def __init__(self, rows: list[dict[str, object]]) -> None:
        self._rows = rows

    def all(self) -> list[dict[str, object]]:
        return self._rows


class _MappingResult:
    def __init__(self, rows: list[dict[str, object]]) -> None:
        self._rows = rows

    def mappings(self) -> _MappingRows:
        return _MappingRows(self._rows)


class _ScalarRows:
    def __init__(self, rows: list[object]) -> None:
        self._rows = rows

    def all(self) -> list[object]:
        return self._rows


class _ScalarResult:
    def __init__(self, rows: list[object]) -> None:
        self._rows = rows

    def scalars(self) -> _ScalarRows:
        return _ScalarRows(self._rows)


class _ExportDB:
    def __init__(self, run_rows: list[dict[str, object]], trial_rows: list[object]) -> None:
        self._results = [_MappingResult(run_rows), _ScalarResult(trial_rows)]

    async def execute(self, stmt: object) -> object:  # noqa: ARG002
        if not self._results:
            raise AssertionError("Unexpected execute() call with no remaining fake results.")
        return self._results.pop(0)


class PoffenbergerExportServiceTests(IsolatedAsyncioTestCase):
    async def test_build_poffenberger_xlsx_includes_joined_runs_and_trials(self) -> None:
        run_id = uuid.uuid4()
        session_id = uuid.uuid4()
        participant_uuid = uuid.uuid4()
        created_at = datetime(2026, 6, 24, 18, 30, tzinfo=timezone.utc)

        run_row = {
            "run_id": run_id,
            "session_id": session_id,
            "participant_uuid": participant_uuid,
            "participant_number": 17,
            "age_band": "18-24",
            "gender": "Woman",
            "handedness": "Right-handed",
            "session_status": "complete",
            "session_created_at": created_at,
            "session_completed_at": created_at,
            "manifest_json": {"blocks": []},
            "started_at": created_at,
            "completed_at": created_at,
            "is_complete": True,
            "total_practice_trials": 10,
            "total_experimental_trials": 600,
            "lh_lvf_total_trials": 150,
            "lh_lvf_valid_rt_trials": 149,
            "lh_lvf_timeout_trials": 1,
            "lh_lvf_invalid_trials": 0,
            "lh_lvf_accurate_trials": 148,
            "lh_lvf_accuracy": Decimal("0.9867"),
            "lh_lvf_mean_rt_ms": Decimal("301.25"),
            "lh_lvf_median_rt_ms": Decimal("299.50"),
            "lh_lvf_sd_rt_ms": Decimal("22.10"),
            "lh_rvf_total_trials": 150,
            "lh_rvf_valid_rt_trials": 150,
            "lh_rvf_timeout_trials": 0,
            "lh_rvf_invalid_trials": 0,
            "lh_rvf_accurate_trials": 147,
            "lh_rvf_accuracy": Decimal("0.9800"),
            "lh_rvf_mean_rt_ms": Decimal("305.25"),
            "lh_rvf_median_rt_ms": Decimal("303.50"),
            "lh_rvf_sd_rt_ms": Decimal("24.10"),
            "rh_lvf_total_trials": 150,
            "rh_lvf_valid_rt_trials": 150,
            "rh_lvf_timeout_trials": 0,
            "rh_lvf_invalid_trials": 0,
            "rh_lvf_accurate_trials": 146,
            "rh_lvf_accuracy": Decimal("0.9733"),
            "rh_lvf_mean_rt_ms": Decimal("306.25"),
            "rh_lvf_median_rt_ms": Decimal("304.50"),
            "rh_lvf_sd_rt_ms": Decimal("25.10"),
            "rh_rvf_total_trials": 150,
            "rh_rvf_valid_rt_trials": 150,
            "rh_rvf_timeout_trials": 0,
            "rh_rvf_invalid_trials": 0,
            "rh_rvf_accurate_trials": 149,
            "rh_rvf_accuracy": Decimal("0.9933"),
            "rh_rvf_mean_rt_ms": Decimal("300.25"),
            "rh_rvf_median_rt_ms": Decimal("298.50"),
            "rh_rvf_sd_rt_ms": Decimal("21.10"),
            "mean_rt_crossed_ms": Decimal("305.75"),
            "mean_rt_uncrossed_ms": Decimal("300.75"),
            "ihtt_difference_ms": Decimal("5.00"),
            "accuracy_crossed": Decimal("0.9767"),
            "accuracy_uncrossed": Decimal("0.9900"),
        }
        trial = SimpleNamespace(
            trial_id=uuid.uuid4(),
            run_id=run_id,
            session_id=session_id,
            participant_uuid=participant_uuid,
            block_number=1,
            trial_number=1,
            global_trial_number=11,
            response_hand="left",
            visual_field="rvf",
            condition_key="lh_rvf",
            is_practice=False,
            is_scored=True,
            expected_key="f",
            pressed_key="f",
            reaction_time_ms=312,
            is_valid_response=True,
            is_timeout=False,
            is_accurate=True,
            jitter_ms=1300,
            client_trial_started_at_ms=Decimal("1000.123"),
            client_stimulus_onset_ms=Decimal("2300.123"),
            client_response_at_ms=Decimal("2612.123"),
            client_trial_ended_at_ms=Decimal("2700.123"),
            created_at=created_at,
        )

        workbook_bytes = await build_poffenberger_xlsx(
            _ExportDB([run_row], [trial]),
            export_date="2026-06-24",
        )

        workbook = openpyxl.load_workbook(io.BytesIO(workbook_bytes))
        assert workbook.sheetnames == ["README", "poffenberger_runs", "poffenberger_trials"]

        runs = workbook["poffenberger_runs"]
        run_headers = [cell.value for cell in runs[1]]
        run_values = {
            header: runs.cell(row=2, column=index + 1).value
            for index, header in enumerate(run_headers)
        }
        assert run_values["participant_number"] == 17
        assert run_values["handedness"] == "Right-handed"
        assert run_values["session_status"] == "complete"
        assert run_values["manifest_json"] == '{"blocks": []}'
        assert run_values["ihtt_difference_ms"] == 5

        trials = workbook["poffenberger_trials"]
        trial_headers = [cell.value for cell in trials[1]]
        trial_values = {
            header: trials.cell(row=2, column=index + 1).value
            for index, header in enumerate(trial_headers)
        }
        assert trial_values["run_id"] == str(run_id)
        assert trial_values["global_trial_number"] == 11
        assert trial_values["condition_key"] == "lh_rvf"
        assert trial_values["reaction_time_ms"] == 312

    async def test_build_sample_poffenberger_xlsx_uses_hardcoded_rows(self) -> None:
        workbook_bytes = build_sample_poffenberger_xlsx(export_date="2026-06-24")

        workbook = openpyxl.load_workbook(io.BytesIO(workbook_bytes))
        readme_values = [
            workbook["README"].cell(row=row_num, column=1).value
            for row_num in range(1, 20)
        ]
        assert any(
            value and "hardcoded sample data" in str(value)
            for value in readme_values
        )

        runs = workbook["poffenberger_runs"]
        run_headers = [cell.value for cell in runs[1]]
        run_values = {
            header: runs.cell(row=2, column=index + 1).value
            for index, header in enumerate(run_headers)
        }
        assert run_values["participant_number"] == 9001
        assert run_values["ihtt_difference_ms"] == 10

        trials = workbook["poffenberger_trials"]
        assert trials.max_row == 6
